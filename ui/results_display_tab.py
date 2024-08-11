from PySide6.QtWidgets import QWidget, QVBoxLayout, QTabWidget, QTableWidget, QTableWidgetItem, QHBoxLayout, \
    QPushButton, QFileDialog
from PySide6.QtCharts import QChart, QChartView, QLineSeries, QScatterSeries, QValueAxis
from PySide6.QtCore import Qt
from PySide6.QtGui import QColor, QPen
import openpyxl
from openpyxl.chart.marker import Marker, DataPoint
from openpyxl import Workbook
from openpyxl.chart import ScatterChart, Reference, Series
from openpyxl.drawing.fill import ColorChoice
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from PySide6.QtWidgets import QFileDialog


class ResultsDisplayTab(QWidget):
    def __init__(self):
        super().__init__()
        self.layout = QVBoxLayout(self)
        self.tab_widget = QTabWidget()
        self.layout.addWidget(self.tab_widget)

        self.nceer_tab = QWidget()
        self.japanese_tab = QWidget()
        self.tab_widget.addTab(self.nceer_tab, "NCEER Results")
        self.tab_widget.addTab(self.japanese_tab, "Japanese Method Results")

        self.setup_nceer_tab()
        self.setup_japanese_tab()

    def setup_nceer_tab(self):
        layout = QVBoxLayout(self.nceer_tab)
        table_chart_layout = QHBoxLayout()
        self.nceer_table = QTableWidget()
        self.nceer_chart_view = QChartView()
        table_chart_layout.addWidget(self.nceer_table)
        table_chart_layout.addWidget(self.nceer_chart_view)
        layout.addLayout(table_chart_layout)
        self.add_save_button(layout, self.nceer_table)

    def setup_japanese_tab(self):
        layout = QVBoxLayout(self.japanese_tab)
        table_chart_layout = QHBoxLayout()
        self.japanese_table = QTableWidget()
        self.japanese_chart_view = QChartView()
        table_chart_layout.addWidget(self.japanese_table)
        table_chart_layout.addWidget(self.japanese_chart_view)
        layout.addLayout(table_chart_layout)
        self.add_save_button(layout, self.japanese_table)

    def update_results(self, nceer_params, japanese_params):
        self.update_nceer_results(nceer_params)
        self.update_japanese_results(japanese_params)

    def update_nceer_results(self, params):
        self.nceer_table.setColumnCount(len(params))
        self.nceer_table.setHorizontalHeaderLabels([
            "Depth", "Total Stress", "Effective Stress", "rd", "CSR", "Cb", "Cs", "Cr", "Cn", "Ce", "N1 60", "\u03B1",
            "\u03B2",
            "N1 60 cs", "CRR 7.5", "MSF", "K\u03B1", "k\u03C3", "CRR", "Fl"
        ])
        self.nceer_table.setRowCount(len(params[0]))

        for row, (
                depth, total_stress, effective_stress, rd, csr, Cb, Cs, Cr, Cn, Ce, n1_60, alpha, beta, n1_60_cs,
                crr_7_5, msf, k_alpha, k_sigma, crr,
                fl) in enumerate(
            zip(*params)):
            items = [
                QTableWidgetItem(f"{depth:.2f}"),
                QTableWidgetItem(f"{total_stress:.2f}"),
                QTableWidgetItem(f"{effective_stress:.2f}"),
                QTableWidgetItem(f"{rd:.3f}"),
                QTableWidgetItem(f"{csr:.4f}"),
                QTableWidgetItem(f"{Cb:.3f}"),
                QTableWidgetItem(f"{Cs:.3f}"),
                QTableWidgetItem(f"{Cr:.3f}"),
                QTableWidgetItem(f"{Cn:.3f}"),
                QTableWidgetItem(f"{Ce:.3f}"),
                QTableWidgetItem(f"{n1_60:.2f}"),
                QTableWidgetItem(f"{alpha:.3f}"),
                QTableWidgetItem(f"{beta:.3f}"),
                QTableWidgetItem(f"{n1_60_cs:.2f}"),
                QTableWidgetItem(f"{crr_7_5:.4f}"),
                QTableWidgetItem(f"{msf:.3f}"),
                QTableWidgetItem(f"{k_alpha:.3f}"),
                QTableWidgetItem(f"{k_sigma:.3f}"),
                QTableWidgetItem(f"{crr:.4f}"),
                QTableWidgetItem(f"{fl:.4f}")
            ]

            for item in items:
                item.setFlags(item.flags() & ~Qt.ItemIsEditable)  # Make the item uneditable

            for col, item in enumerate(items):
                self.nceer_table.setItem(row, col, item)

        self.update_chart(self.nceer_chart_view, params[0], params[len(params) - 1], "Fl vs Depth")

    def update_japanese_results(self, params):
        self.japanese_table.setColumnCount(len(params))
        self.japanese_table.setHorizontalHeaderLabels([
            "Depth", "Total Stress", "Effective Stress", "rd", "CSR", "N1", "a", "b", "Na", "Rl 7.5", "MSF", "K\u03B1",
            "k\u03C3", "Rl", "Fl"
        ])
        self.japanese_table.setRowCount(len(params[0]))

        for row, (
                depth, total_stress, effective_stress, rd, csr, n1_60, a, b, n1_60_cs, crr_7_5, msf, k_alpha, k_sigma,
                rl,
                fl) in enumerate(
            zip(*params)):
            items = [
                QTableWidgetItem(f"{depth:.2f}"),
                QTableWidgetItem(f"{total_stress:.2f}"),
                QTableWidgetItem(f"{effective_stress:.2f}"),
                QTableWidgetItem(f"{rd:.3f}"),
                QTableWidgetItem(f"{csr:.4f}"),
                QTableWidgetItem(f"{n1_60:.2f}"),
                QTableWidgetItem(f"{a:.3f}"),
                QTableWidgetItem(f"{b:.3f}"),
                QTableWidgetItem(f"{n1_60_cs:.2f}"),
                QTableWidgetItem(f"{crr_7_5:.4f}"),
                QTableWidgetItem(f"{msf:.3f}"),
                QTableWidgetItem(f"{k_alpha:.3f}"),
                QTableWidgetItem(f"{k_sigma:.3f}"),
                QTableWidgetItem(f"{rl:.4f}"),
                QTableWidgetItem(f"{fl:.4f}")
            ]

            for item in items:
                item.setFlags(item.flags() & ~Qt.ItemIsEditable)  # Make the item uneditable

            for col, item in enumerate(items):
                self.japanese_table.setItem(row, col, item)

        self.update_chart(self.japanese_chart_view, params[0], params[len(params) - 1], "Fl vs Depth")

    def update_chart(self, chart_view, depths, fls, title):
        line_series = QLineSeries()
        scatter_series = QScatterSeries()
        scatter_series2 = QScatterSeries()

        for depth, fl in zip(depths, fls):
            if fl < 1:
                scatter_series.append(fl, -depth)
            else:
                scatter_series2.append(fl, -depth)

            line_series.append(fl, -depth)  # Negative depth to show increasing depth downwards

        chart = QChart()
        chart.addSeries(line_series)
        chart.addSeries(scatter_series)
        chart.addSeries(scatter_series2)
        chart.setTitle(title)

        # Customize appearance
        line_color = QColor(Qt.black)
        scatter_color = QColor(Qt.red)
        scatter_color2 = QColor(Qt.green)

        line_series.setColor(line_color)
        line_series.setPen(QPen(line_color, 2))

        scatter_series.setColor(scatter_color)
        scatter_series.setMarkerSize(10)
        scatter_series.setBorderColor(scatter_color)
        scatter_series.setBrush(Qt.red)  # This makes the points "clear" (white fill)

        scatter_series2.setColor(scatter_color2)
        scatter_series2.setMarkerSize(10)
        scatter_series2.setBorderColor(scatter_color2)
        scatter_series2.setBrush(Qt.green)  # This makes the points "clear" (white fill)

        axis_x = QValueAxis()
        axis_x.setTitleText("Fl")
        chart.addAxis(axis_x, Qt.AlignBottom)
        line_series.attachAxis(axis_x)
        scatter_series.attachAxis(axis_x)
        scatter_series2.attachAxis(axis_x)

        axis_y = QValueAxis()
        axis_y.setTitleText("Depth (m)")
        chart.addAxis(axis_y, Qt.AlignLeft)
        line_series.attachAxis(axis_y)
        scatter_series.attachAxis(axis_y)
        scatter_series2.attachAxis(axis_y)
        # Adjust the range to show all points
        min_fl = min(fls)
        max_fl = max(fls)
        min_depth = min(depths)
        max_depth = max(depths)

        axis_x.setRange(min_fl * 0.9, max_fl * 1.1)
        axis_y.setRange(-max_depth * 1.1, -min_depth * 0.9)

        chart_view.setChart(chart)

    def add_save_button(self, layout, table):
        save_button = QPushButton("Save as XLSX")
        save_button.clicked.connect(lambda: self.save_to_xlsx(table))

        layout.addWidget(save_button)

    def save_to_xlsx(self, table):
        file_path, _ = QFileDialog.getSaveFileName(self, "Save Excel File", "", "Excel Files (*.xlsx)")
        if not file_path:
            return

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Data"

        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                             top=Side(style='thin'), bottom=Side(style='thin'))
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

        # Write headers
        for col in range(table.columnCount()):
            cell = sheet.cell(row=1, column=col + 1, value=table.horizontalHeaderItem(col).text())
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border

        # Write data and find Depth and Fl columns
        depth_col = fl_col = None
        max_depth = 0
        max_fl = 0

        # Separate rows into two groups based on Fl value
        rows_red = []
        rows_green = []

        for row in range(table.rowCount()):
            row_data = []
            for col in range(table.columnCount()):
                item = table.item(row, col)
                if item is not None:
                    cell = sheet.cell(row=row + 2, column=col + 1)
                    try:
                        value = float(item.text())
                        cell.value = value

                        if table.horizontalHeaderItem(col).text() == "Depth":
                            depth_col = col + 1
                            max_depth = max(max_depth, value)
                        elif table.horizontalHeaderItem(col).text() == "Fl":
                            fl_col = col + 1
                            max_fl = max(max_fl, value)
                            if value < 1:
                                cell.fill = red_fill
                                rows_red.append(row + 2)
                            else:
                                cell.fill = green_fill
                                rows_green.append(row + 2)
                    except ValueError:
                        cell.value = item.text()

        if depth_col is None or fl_col is None:
            print("Error: Couldn't find Depth or Fl column")
            return

        # Create chart
        chart = ScatterChart()
        chart.title = "Depth vs Fl"
        chart.x_axis.title = "Fl"
        chart.y_axis.title = "Depth (m)"

        # Function to create a series from the given rows
        def create_series(rows, color, title):
            xvalues = Reference(sheet, min_col=fl_col, min_row=min(rows), max_row=max(rows))
            yvalues = Reference(sheet, min_col=depth_col, min_row=min(rows), max_row=max(rows))
            series = Series(values=yvalues, xvalues=xvalues, title=title)
            series.marker = Marker(symbol="circle", size=10)
            series.marker.graphicalProperties.solidFill = ColorChoice(prstClr=color)
            series.graphicalProperties.line.solidFill = "000000"  # Black line to connect points
            return series

        # Create red and green series
        if rows_red:
            red_series = create_series(rows_red, "red", "Fl < 1")
            chart.series.append(red_series)

        if rows_green:
            green_series = create_series(rows_green, "green", "Fl >= 1")
            chart.series.append(green_series)

        # Customize chart
        chart.x_axis.scaling.min = 0
        chart.x_axis.scaling.max = max_fl * 1.1  # Slightly increase the max to give space
        chart.y_axis.scaling.min = 0
        chart.y_axis.scaling.max = max_depth * 1.1  # Add 10% to the max depth for better visibility

        # Set chart size
        chart.height = 15  # Height in cm
        chart.width = 20  # Width in cm

        # Add chart to a new sheet
        chart_sheet = workbook.create_sheet(title="Depth vs Fl Chart")
        chart_sheet.add_chart(chart, "A1")

        # Save workbook
        workbook.save(file_path)
        print(f"Excel file saved successfully at {file_path}")
